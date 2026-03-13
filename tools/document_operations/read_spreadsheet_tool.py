from __future__ import annotations

import csv
from pathlib import Path
from typing import Any, Dict, List, Optional

from protocol_monk.exceptions.tools import ToolError
from protocol_monk.tools.base import BaseTool
from protocol_monk.tools.output_contract import build_range_pagination, build_tool_output


class ReadSpreadsheetTool(BaseTool):
    """Read CSV and Excel workbooks as structured row slices."""

    MAX_FILE_SIZE_BYTES = 8 * 1024 * 1024
    DEFAULT_ROW_LIMIT = 100

    @property
    def name(self) -> str:
        return "read_spreadsheet"

    @property
    def description(self) -> str:
        return (
            "Read CSV or Excel spreadsheets as structured columns and row slices. "
            "Defaults to the first 100 rows."
        )

    @property
    def parameter_schema(self) -> Dict[str, Any]:
        return {
            "type": "object",
            "properties": {
                "filepath": {"type": "string", "description": "Spreadsheet path"},
                "sheet_name": {
                    "type": "string",
                    "description": "Optional sheet name for Excel workbooks",
                },
                "row_start": {
                    "type": "integer",
                    "description": "Starting row number (1-based, inclusive)",
                    "default": 1,
                },
                "row_limit": {
                    "type": "integer",
                    "description": "Maximum number of rows to return",
                    "default": self.DEFAULT_ROW_LIMIT,
                },
            },
            "required": ["filepath"],
        }

    async def run(self, **kwargs) -> Dict[str, Any]:
        return self._execute_sync(**kwargs)

    def _execute_sync(self, **kwargs) -> Dict[str, Any]:
        filepath = kwargs.get("filepath")
        if not filepath:
            raise ToolError(
                "Missing required parameter: 'filepath'",
                user_hint="Please provide a spreadsheet filepath.",
            )

        row_start = int(kwargs.get("row_start", 1) or 1)
        row_limit = int(
            kwargs.get("row_limit", self.DEFAULT_ROW_LIMIT) or self.DEFAULT_ROW_LIMIT
        )
        sheet_name = kwargs.get("sheet_name")
        if row_start < 1:
            raise ToolError(
                "row_start must be >= 1",
                user_hint="Spreadsheet row_start must be 1 or greater.",
            )
        if row_limit < 1:
            raise ToolError(
                "row_limit must be >= 1",
                user_hint="Spreadsheet row_limit must be 1 or greater.",
            )

        full_path = self.path_validator.validate_path(filepath, must_exist=False)
        self._validate_file(full_path)

        suffix = full_path.suffix.lower()
        if suffix == ".numbers":
            raise ToolError(
                "Apple Numbers files are not supported in this pass.",
                user_hint="Numbers support is deferred. Use CSV or Excel for now.",
                details={"path": str(full_path)},
            )

        if suffix == ".csv":
            workbook = self._read_csv(full_path)
        elif suffix == ".xlsx":
            workbook = self._read_xlsx(full_path, sheet_name=sheet_name)
        elif suffix == ".xls":
            workbook = self._read_xls(full_path, sheet_name=sheet_name)
        else:
            raise ToolError(
                f"Unsupported spreadsheet format: {suffix or 'unknown'}",
                user_hint="Supported spreadsheet formats are .csv, .xlsx, and .xls.",
                details={"path": str(full_path), "suffix": suffix},
            )

        total_rows = workbook["total_rows"]
        total_columns = workbook["total_columns"]
        rows = workbook["rows"]
        actual_start = row_start
        actual_end = min(total_rows, row_start + row_limit - 1)
        if total_rows == 0:
            actual_start = 1
            actual_end = 0
        elif row_start > total_rows:
            raise ToolError(
                f"Requested row_start {row_start} exceeds total rows {total_rows}.",
                user_hint=f"Requested row_start {row_start} is beyond the sheet length.",
                details={"row_start": row_start, "total_rows": total_rows},
            )

        sliced_rows = rows[row_start - 1 : actual_end]
        row_records = [
            {
                "row_number": row_start + index,
                "values": self._pad_row(row, total_columns),
            }
            for index, row in enumerate(sliced_rows)
        ]

        pagination = build_range_pagination(
            mode="row_range",
            total_items=total_rows,
            returned_start=actual_start,
            returned_end=actual_end,
            page_size=max(1, row_limit),
            start_key="row_start",
            end_key="row_end",
            total_key="total_rows",
        )

        selected_sheet = workbook["selected_sheet"]
        summary_target = selected_sheet or full_path.name
        if actual_end >= actual_start:
            summary = f"Read rows {actual_start}-{actual_end} from {summary_target}."
        else:
            summary = f"Read 0 rows from {summary_target}."

        return build_tool_output(
            result_type="spreadsheet_read",
            summary=summary,
            data={
                "filepath": str(full_path),
                "format": suffix.lstrip("."),
                "available_sheets": workbook["available_sheets"],
                "selected_sheet": selected_sheet,
                "total_rows": total_rows,
                "total_columns": total_columns,
                "columns": self._build_columns(rows, total_columns),
                "rows": row_records,
            },
            pagination=pagination,
        )

    def _validate_file(self, path: Path) -> None:
        try:
            stat = path.stat()
        except FileNotFoundError:
            raise ToolError(
                f"Spreadsheet not found: {path.name}",
                user_hint=f"The file '{path.name}' does not exist.",
                details={"path": str(path)},
            )
        except PermissionError:
            raise ToolError(
                f"Permission denied: {path.name}",
                user_hint=f"No permission to read '{path.name}'.",
                details={"path": str(path)},
            )

        if stat.st_size > self.MAX_FILE_SIZE_BYTES:
            raise ToolError(
                "Spreadsheet is too large to load in one call.",
                user_hint="Spreadsheet is too large for the document reader.",
                details={
                    "path": str(path),
                    "actual_size_bytes": stat.st_size,
                    "max_size_bytes": self.MAX_FILE_SIZE_BYTES,
                },
            )

    def _read_csv(self, path: Path) -> Dict[str, Any]:
        with path.open("r", encoding="utf-8", newline="") as handle:
            reader = csv.reader(handle)
            rows = [[self._normalize_cell(cell) for cell in row] for row in reader]
        return {
            "available_sheets": ["Sheet1"],
            "selected_sheet": "Sheet1",
            "rows": rows,
            "total_rows": len(rows),
            "total_columns": max((len(row) for row in rows), default=0),
        }

    def _read_xlsx(self, path: Path, *, sheet_name: Optional[str]) -> Dict[str, Any]:
        try:
            from openpyxl import load_workbook
        except Exception as exc:  # pragma: no cover
            raise ToolError(
                "openpyxl is not installed.",
                user_hint="Install openpyxl to read .xlsx files.",
                details={"error": str(exc)},
            )

        workbook = load_workbook(path, data_only=True, read_only=True)
        try:
            available_sheets = list(workbook.sheetnames)
            selected_sheet = sheet_name or (available_sheets[0] if available_sheets else None)
            if selected_sheet not in workbook.sheetnames:
                raise ToolError(
                    f"Sheet '{selected_sheet}' does not exist.",
                    user_hint=f"Available sheets: {', '.join(available_sheets)}",
                    details={
                        "sheet_name": selected_sheet,
                        "available_sheets": available_sheets,
                    },
                )

            worksheet = workbook[selected_sheet]
            rows = [
                [self._normalize_cell(cell) for cell in row]
                for row in worksheet.iter_rows(values_only=True)
            ]
            return {
                "available_sheets": available_sheets,
                "selected_sheet": selected_sheet,
                "rows": rows,
                "total_rows": len(rows),
                "total_columns": max((len(row) for row in rows), default=0),
            }
        finally:
            workbook.close()

    def _read_xls(self, path: Path, *, sheet_name: Optional[str]) -> Dict[str, Any]:
        try:
            import xlrd
        except Exception as exc:  # pragma: no cover
            raise ToolError(
                "xlrd is not installed.",
                user_hint="Install xlrd to read .xls files.",
                details={"error": str(exc)},
            )

        workbook = xlrd.open_workbook(path)
        available_sheets = workbook.sheet_names()
        selected_sheet = sheet_name or (available_sheets[0] if available_sheets else None)
        if selected_sheet not in available_sheets:
            raise ToolError(
                f"Sheet '{selected_sheet}' does not exist.",
                user_hint=f"Available sheets: {', '.join(available_sheets)}",
                details={"sheet_name": selected_sheet, "available_sheets": available_sheets},
            )

        sheet = workbook.sheet_by_name(selected_sheet)
        rows = [
            [self._normalize_cell(sheet.cell_value(rowx, colx)) for colx in range(sheet.ncols)]
            for rowx in range(sheet.nrows)
        ]
        return {
            "available_sheets": available_sheets,
            "selected_sheet": selected_sheet,
            "rows": rows,
            "total_rows": len(rows),
            "total_columns": max((len(row) for row in rows), default=0),
        }

    def _normalize_cell(self, value: Any) -> Any:
        if value is None:
            return None
        if isinstance(value, bool):
            return value
        if isinstance(value, (int, float)):
            return value
        text = str(value)
        return text if text != "" else None

    def _pad_row(self, row: List[Any], total_columns: int) -> List[Any]:
        if len(row) >= total_columns:
            return row
        return row + [None] * (total_columns - len(row))

    def _build_columns(
        self,
        rows: List[List[Any]],
        total_columns: int,
    ) -> List[Dict[str, Any]]:
        headers = rows[0] if rows else []
        columns = []
        for index in range(total_columns):
            label = None
            if index < len(headers) and headers[index] not in (None, ""):
                label = str(headers[index])
            if not label:
                label = self._excel_column_label(index + 1)
            columns.append({"index": index + 1, "label": label})
        return columns

    def _excel_column_label(self, index: int) -> str:
        label = ""
        current = index
        while current > 0:
            current, remainder = divmod(current - 1, 26)
            label = chr(65 + remainder) + label
        return label
